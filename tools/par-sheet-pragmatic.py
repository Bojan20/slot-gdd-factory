#!/usr/bin/env python3
"""tools/par-sheet-pragmatic.py

MATH-DEEP HYB-4 (2026-06-23) — Pragmatic Play PAR sheet xlsx adapter.

Vendor-specific layout
  Pragmatic xlsx exports use Spanish column headers (legacy from RSI:
  the underlying math toolchain was originally developed in Spain):
    "Rodillo 1" .. "Rodillo 5"   (Reel 1..5)
    "Símbolo"                    (Symbol)
    "Probabilidad" / "Frecuencia" (Probability / Frequency)
    "Pago"                       (Pay)
  Many modern Pragmatic exports ALSO include English aliases — adapter
  handles both. Sheet names commonly start with "Carrete" (single reel
  detail) or "Sheet1" (combined paytable).

What the adapter emits
  Identical canonical ParSheet shape as par-sheet-xlsx-ingest.py:
    {
      "vendor": "pragmatic",
      "reels": [[id, id, ...], ...],          # expanded per weight
      "per_reel_weights": { "0": { id: w, ... }, ... },
      "paytable": [{ "symbolId": id, "combos": { "3": x, "4": y, "5": z } }],
      "totals": { "reels": 5, "symbols": N, "sumWeight": ... }
    }

USAGE
  python3 tools/par-sheet-pragmatic.py --xlsx <path> [--sheet <name>] [--out -]

EXIT
  0 — extraction successful
  1 — file missing OR openpyxl not installed
  2 — sheet not found / required columns missing

HARD RULE #1 (vendor-neutral output)
  Vendor identifier 'pragmatic' is a routing code, not a product name.
  All symbol ids + paytable values flow through canonical schema.
"""
import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl-missing — pip3 install openpyxl"}))
    sys.exit(1)


# Pragmatic header recognizers (Spanish primary, English alias).
SYMBOL_HEADERS = {
    "símbolo", "simbolo", "symbol", "sym", "name", "símb.",
}
REEL_HEADER_PREFIXES_ES = {"rodillo", "carrete"}
REEL_HEADER_PREFIXES_EN = {"reel", "strip"}
PAY_HEADER_PATTERNS = {
    "pago", "pay", "payout",
    "3 oak", "4 oak", "5 oak", "3 of a kind", "4 of a kind", "5 of a kind",
}


def _norm_header(s):
    """Lowercase + strip; tolerates accents (NFKD normalize)."""
    import unicodedata
    if s is None:
        return ""
    out = unicodedata.normalize("NFKD", str(s)).strip().lower()
    # Strip combining marks (accents) for header matching.
    return "".join(ch for ch in out if not unicodedata.combining(ch))


def _is_reel_header(s):
    n = _norm_header(s)
    for prefix in REEL_HEADER_PREFIXES_ES | REEL_HEADER_PREFIXES_EN:
        if n.startswith(prefix):
            # Must end with a digit (Rodillo 1, Reel 1, Strip 5 etc).
            rest = n[len(prefix):].strip()
            if rest and rest.replace(" ", "").isdigit():
                return True
    return False


def _is_pay_header(s):
    n = _norm_header(s)
    if n in PAY_HEADER_PATTERNS:
        return True
    # "3 oak", "4-oak", "5oak", "pay3", "pago 3" forms
    import re
    if re.match(r"^(pay|pago)\s*\d+$", n):
        return True
    if re.match(r"^\d+\s*(oak|de\s+un\s+tipo|of\s+a\s+kind)$", n):
        return True
    return False


def _pay_count(header):
    """Return the N-of-a-kind count from a pay header (3, 4, 5)."""
    import re
    n = _norm_header(header)
    m = re.search(r"\d+", n)
    return m.group(0) if m else None


def extract(xlsx_path, sheet_name=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheets = wb.sheetnames
    # Auto-detect: prefer sheet whose name starts with "Carrete" or "Sheet1"
    # OR sheet that contains the most Rodillo headers in row 1.
    candidates = []
    target = sheet_name
    if target and target not in sheets:
        print(json.dumps({"error": f"sheet {target!r} not found; available: {sheets}"}))
        sys.exit(2)
    if not target:
        # Score each sheet by reel-header count in first few rows.
        best_score = -1
        best_sheet = sheets[0]
        for sn in sheets[:5]:
            ws = wb[sn]
            score = 0
            for r in range(1, min(6, ws.max_row + 1)):
                for c in range(1, min(15, ws.max_column + 1)):
                    v = ws.cell(row=r, column=c).value
                    if _is_reel_header(v):
                        score += 1
            candidates.append((sn, score))
            if score > best_score:
                best_score = score
                best_sheet = sn
        target = best_sheet
    ws = wb[target]

    # Find the header row: row with ≥1 symbol-header AND ≥1 reel-header.
    header_row = None
    for r in range(1, min(30, ws.max_row + 1)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        has_sym = any(_norm_header(v) in SYMBOL_HEADERS for v in row_vals)
        has_reel = any(_is_reel_header(v) for v in row_vals)
        if has_sym and has_reel:
            header_row = r
            break
    if header_row is None:
        print(json.dumps({"error": f"no header row in sheet {target} (need both Símbolo + Rodillo cols)"}))
        sys.exit(2)

    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    sym_col = None
    reel_cols = []
    pay_cols = {}
    for i, h in enumerate(headers):
        if _norm_header(h) in SYMBOL_HEADERS:
            if sym_col is None:
                sym_col = i + 1
        elif _is_reel_header(h):
            reel_cols.append(i + 1)
        elif _is_pay_header(h):
            cnt = _pay_count(h)
            if cnt:
                pay_cols[cnt] = i + 1

    if sym_col is None or not reel_cols:
        print(json.dumps({"error": "header detection failed (no symbol col OR no reel cols)"}))
        sys.exit(2)

    symbols = []
    weights_by_reel = [[] for _ in reel_cols]
    pay_rows = []

    for r in range(header_row + 1, ws.max_row + 1):
        sym_raw = ws.cell(row=r, column=sym_col).value
        if sym_raw is None or str(sym_raw).strip() == "":
            continue
        sym = str(sym_raw).strip()
        symbols.append(sym)
        for k, rc in enumerate(reel_cols):
            v = ws.cell(row=r, column=rc).value
            try:
                w = int(v) if v is not None else 0
            except (ValueError, TypeError):
                w = 0
            # UQ-DEEP-C audit fix (D-HIGH-negative): negative weights
            # corrupt pHit() math (perReelSum can become 0 or negative),
            # producing NaN/Infinity RTP. Clamp to 0 — treat negative
            # values as malformed source data, not legitimate weights.
            if w < 0:
                w = 0
            weights_by_reel[k].append(w)
        combos = {}
        for cnt, pc in pay_cols.items():
            v = ws.cell(row=r, column=pc).value
            try:
                pay = int(v) if v is not None else 0
            except (ValueError, TypeError):
                pay = 0
            if pay > 0:
                combos[cnt] = pay
        if combos:
            pay_rows.append({"symbolId": sym, "combos": combos})

    # Expand reels by weight.
    # UQ-DEEP-C audit fix (D-MED-reels-dos): cap per-reel expansion at
    # 500k elements so a malformed PAR sheet with weight=10_000_000
    # can't balloon the strip into hundreds of MB. Real PARs use
    # ≤ 200k per reel — 500k is a generous safety ceiling.
    MAX_REEL_EXPANSION = 500_000
    reels = []
    for w_arr in weights_by_reel:
        strip = []
        for i, s in enumerate(symbols):
            remaining = MAX_REEL_EXPANSION - len(strip)
            if remaining <= 0:
                break
            count = min(max(0, w_arr[i]), remaining)
            strip.extend([s] * count)
        reels.append(strip)
    per_reel_weights = {}
    for idx, w_arr in enumerate(weights_by_reel):
        m = {}
        for i, s in enumerate(symbols):
            if w_arr[i] > 0:
                m[s] = w_arr[i]
        per_reel_weights[str(idx)] = m
    sum_weight = sum(sum(w) for w in weights_by_reel)
    per_reel_sum = [sum(w) for w in weights_by_reel]

    return {
        "vendor": "pragmatic",
        "sheet_used": target,
        "header_row": header_row,
        "reels": reels,
        "per_reel_weights": per_reel_weights,
        "paytable": pay_rows,
        "totals": {
            "reels": len(reel_cols),
            "symbols": len(symbols),
            "sumWeight": sum_weight,
            "perReelSum": per_reel_sum,
        },
    }


def main():
    ap = argparse.ArgumentParser(description="Pragmatic PAR sheet xlsx adapter")
    ap.add_argument("--xlsx", required=True, help="Path to .xlsx file")
    ap.add_argument("--sheet", help="Sheet name (auto-detect if omitted)")
    ap.add_argument("--out", default="-", help="Output path or '-' for stdout")
    args = ap.parse_args()

    # UQ-DEEP-C audit fix (D-HIGH-sheet-flag-injection): even though
    # spawnSync's array form is shell-injection-safe, a sheet name like
    # "--version" or "-h" would otherwise be re-interpreted as a flag
    # by argparse if a future caller switches to single-arg invocation.
    # Lock the input shape to a strict whitelist matching what the
    # par-sheet-detect.mjs regex already enforces.
    import re as _re
    if args.sheet is not None:
        if not _re.match(r"^[A-Za-z0-9._-]{1,80}$", args.sheet):
            print(json.dumps({"error": f"invalid --sheet name: {args.sheet!r}"}))
            sys.exit(2)

    p = Path(args.xlsx)
    if not p.exists():
        print(json.dumps({"error": f"file not found: {p}"}))
        sys.exit(1)

    try:
        result = extract(str(p), args.sheet)
    except Exception as e:
        print(json.dumps({"error": f"extraction failed: {e}"}))
        sys.exit(2)

    out = json.dumps(result, indent=2)
    if args.out == "-" or not args.out:
        print(out)
    else:
        Path(args.out).write_text(out)
        print(f"▸ wrote {args.out}", file=sys.stderr)


if __name__ == "__main__":
    main()
