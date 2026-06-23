#!/usr/bin/env python3
"""tools/par-sheet-lw.py

MATH-DEEP HYB-4 (2026-06-23) — Light & Wonder (formerly Scientific Games)
PAR sheet xlsx adapter.

Vendor-specific layout
  L&W xlsx exports inherit Bally/SG Bryt math-toolchain layout. Distinct
  signature:
    "STRIP_1" .. "STRIP_5"   OR   "Strip 1" .. "Strip 5"   (reel columns)
    "SYM" or "Symbol"                                        (symbol id)
    "Weight" / "W"                                            (alt to numeric strip)
    "3OAK" / "4OAK" / "5OAK"                                  (pay columns)
  Sheet names usually start with "Brytt" or "Strips". Some exports use
  uppercase-only headers (legacy mainframe export).

What the adapter emits
  Canonical ParSheet shape (vendor: "lw"):
    {
      "vendor": "lw",
      "reels": [[id, ...], ...],
      "per_reel_weights": { "0": { id: w, ... }, ... },
      "paytable": [{ "symbolId": id, "combos": { "3": x, "4": y, "5": z } }],
      "totals": { ... }
    }

USAGE
  python3 tools/par-sheet-lw.py --xlsx <path> [--sheet <name>] [--out -]

EXIT
  0 — extraction successful
  1 — file missing OR openpyxl not installed
  2 — sheet not found / required columns missing

HARD RULE #1 (vendor-neutral output)
  Vendor identifier 'lw' is a routing code, not a product name.
"""
import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl-missing — pip3 install openpyxl"}))
    sys.exit(1)


SYMBOL_HEADERS = {"sym", "symbol", "name"}
REEL_HEADER_RE_PATTERNS = [
    r"^strip[_\s-]?\d+$",
    r"^reel[_\s-]?\d+$",
]
PAY_HEADER_RE_PATTERNS = [
    r"^\d+oak$",
    r"^\d+\s*oak$",
    r"^pay\d+$",
    r"^pay\s*\d+$",
]


def _norm(s):
    return "" if s is None else str(s).strip().lower()


def _is_reel_header(s):
    import re
    n = _norm(s)
    for pat in REEL_HEADER_RE_PATTERNS:
        if re.match(pat, n):
            return True
    return False


def _is_pay_header(s):
    import re
    n = _norm(s)
    for pat in PAY_HEADER_RE_PATTERNS:
        if re.match(pat, n):
            return True
    return False


def _pay_count(header):
    import re
    n = _norm(header)
    m = re.search(r"\d+", n)
    return m.group(0) if m else None


def extract(xlsx_path, sheet_name=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheets = wb.sheetnames
    target = sheet_name
    if target and target not in sheets:
        print(json.dumps({"error": f"sheet {target!r} not found; available: {sheets}"}))
        sys.exit(2)
    if not target:
        # Prefer sheet starting with "Brytt" or "Strip"; else score by reel-header count.
        for sn in sheets:
            ln = _norm(sn)
            if ln.startswith("brytt") or ln.startswith("strip"):
                target = sn
                break
        if not target:
            best_score = -1
            best_sheet = sheets[0]
            for sn in sheets[:5]:
                ws = wb[sn]
                score = 0
                for r in range(1, min(6, ws.max_row + 1)):
                    for c in range(1, min(15, ws.max_column + 1)):
                        v = ws.cell(row=r, column=c).value
                        if _is_reel_header(v):
                            score += 1
                if score > best_score:
                    best_score = score
                    best_sheet = sn
            target = best_sheet
    ws = wb[target]

    header_row = None
    for r in range(1, min(30, ws.max_row + 1)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        has_sym = any(_norm(v) in SYMBOL_HEADERS for v in row_vals)
        has_reel = any(_is_reel_header(v) for v in row_vals)
        if has_sym and has_reel:
            header_row = r
            break
    if header_row is None:
        print(json.dumps({"error": f"no header row in sheet {target} (need both SYM + STRIP cols)"}))
        sys.exit(2)

    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    sym_col = None
    reel_cols = []
    pay_cols = {}
    for i, h in enumerate(headers):
        if _norm(h) in SYMBOL_HEADERS:
            if sym_col is None:
                sym_col = i + 1
        elif _is_reel_header(h):
            reel_cols.append(i + 1)
        elif _is_pay_header(h):
            cnt = _pay_count(h)
            if cnt:
                pay_cols[cnt] = i + 1

    if sym_col is None or not reel_cols:
        print(json.dumps({"error": "header detection failed (no SYM col OR no STRIP cols)"}))
        sys.exit(2)

    symbols = []
    weights_by_reel = [[] for _ in reel_cols]
    pay_rows = []

    for r in range(header_row + 1, ws.max_row + 1):
        sym_raw = ws.cell(row=r, column=sym_col).value
        if sym_raw is None or str(sym_raw).strip() == "":
            continue
        sym = str(sym_raw).strip()
        symbols.append(sym)
        for k, rc in enumerate(reel_cols):
            v = ws.cell(row=r, column=rc).value
            try:
                w = int(v) if v is not None else 0
            except (ValueError, TypeError):
                w = 0
            weights_by_reel[k].append(w)
        combos = {}
        for cnt, pc in pay_cols.items():
            v = ws.cell(row=r, column=pc).value
            try:
                pay = int(v) if v is not None else 0
            except (ValueError, TypeError):
                pay = 0
            if pay > 0:
                combos[cnt] = pay
        if combos:
            pay_rows.append({"symbolId": sym, "combos": combos})

    reels = []
    for w_arr in weights_by_reel:
        strip = []
        for i, s in enumerate(symbols):
            strip.extend([s] * w_arr[i])
        reels.append(strip)
    per_reel_weights = {}
    for idx, w_arr in enumerate(weights_by_reel):
        m = {}
        for i, s in enumerate(symbols):
            if w_arr[i] > 0:
                m[s] = w_arr[i]
        per_reel_weights[str(idx)] = m
    sum_weight = sum(sum(w) for w in weights_by_reel)
    per_reel_sum = [sum(w) for w in weights_by_reel]

    return {
        "vendor": "lw",
        "sheet_used": target,
        "header_row": header_row,
        "reels": reels,
        "per_reel_weights": per_reel_weights,
        "paytable": pay_rows,
        "totals": {
            "reels": len(reel_cols),
            "symbols": len(symbols),
            "sumWeight": sum_weight,
            "perReelSum": per_reel_sum,
        },
    }


def main():
    ap = argparse.ArgumentParser(description="L&W / Brytt PAR sheet xlsx adapter")
    ap.add_argument("--xlsx", required=True, help="Path to .xlsx file")
    ap.add_argument("--sheet", help="Sheet name (auto-detect if omitted)")
    ap.add_argument("--out", default="-", help="Output path or '-' for stdout")
    args = ap.parse_args()

    p = Path(args.xlsx)
    if not p.exists():
        print(json.dumps({"error": f"file not found: {p}"}))
        sys.exit(1)

    try:
        result = extract(str(p), args.sheet)
    except Exception as e:
        print(json.dumps({"error": f"extraction failed: {e}"}))
        sys.exit(2)

    out = json.dumps(result, indent=2)
    if args.out == "-" or not args.out:
        print(out)
    else:
        Path(args.out).write_text(out)
        print(f"▸ wrote {args.out}", file=sys.stderr)


if __name__ == "__main__":
    main()
