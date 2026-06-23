#!/usr/bin/env python3
"""tools/par-sheet-xlsx-ingest.py

MATH-PRECISION-4 — Real par sheet ingest from xlsx.

Reads Cash Eruption par sheet (~/Desktop/Bojan/ParSheets_CashEruption 1.xlsx)
and emits JSON sa:
  - per_reel_weights[reelIdx][symbolName] = absolute weight
  - per_reel_totals[reelIdx] = sum (typically 100000)
  - paytable[symbolName][matchCount] = pay × bet
  - rtp_variant_id (e.g. 200-1637-001)

USAGE
  python3 tools/par-sheet-xlsx-ingest.py --xlsx PATH --sheet PAR-001 --out OUT.json

EXIT
  0 — extraction successful
  1 — file missing or shape unexpected
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("▸ openpyxl not installed; run: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


def extract(xlsx_path, sheet_name):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    # UQ-DEEP-L fix (Boki 2026-06-23): Cash-vendor-style XLSX uses default
    # sheet name 'PAR-001'. Vendor-neutral regulator-grade XLSX uses
    # numbered tabs ('05 Reel Composition' / '05 Reel Strip Composition').
    # If --sheet param doesn't match exactly, auto-detect first sheet whose
    # name contains 'Reel' AND 'Comp' (or 'Strip') so the operator doesn't
    # need to manually pass --sheet for every vendor format.
    if sheet_name not in wb.sheetnames:
        import re as _re
        candidates = [
            s for s in wb.sheetnames
            if _re.search(r'\breel.*(?:comp|strip)|composition|strip', s, _re.IGNORECASE)
        ]
        if candidates:
            print(f"▸ sheet {sheet_name!r} not found, auto-selected {candidates[0]!r}", file=sys.stderr)
            sheet_name = candidates[0]
        else:
            print(f"▸ sheet {sheet_name!r} not found; available: {wb.sheetnames}", file=sys.stderr)
            sys.exit(1)
    ws = wb[sheet_name]

    title = ws.cell(row=1, column=1).value
    topology = ws.cell(row=2, column=1).value
    swid_row = ws.cell(row=3, column=5).value

    # Walk rows looking for "Reel 1 | Reel 2 ..." header (col D-H)
    weight_header_row = None
    for r in range(1, 40):
        row_vals = [ws.cell(row=r, column=c).value for c in range(4, 9)]
        if row_vals == ['Reel 1', 'Reel 2', 'Reel 3', 'Reel 4', 'Reel 5']:
            weight_header_row = r
            break

    if weight_header_row is None:
        print("▸ no 'Reel 1..5' header found", file=sys.stderr)
        sys.exit(1)

    # Symbol weights table starts row after header
    per_reel_weights = {f"reel{i}": {} for i in range(5)}
    symbols = []
    r = weight_header_row + 1
    while r < weight_header_row + 30:
        sym = ws.cell(row=r, column=3).value
        if sym in (None, '', 'Total'):
            if sym == 'Total':
                break
            r += 1
            continue
        weights = [ws.cell(row=r, column=4 + i).value for i in range(5)]
        if not any(isinstance(w, (int, float)) for w in weights):
            r += 1
            continue
        for i, w in enumerate(weights):
            if isinstance(w, (int, float)):
                per_reel_weights[f"reel{i}"][sym] = round(float(w), 4)
        symbols.append(sym)
        r += 1

    # Totals row
    per_reel_totals = {}
    total_row = r
    for i in range(5):
        v = ws.cell(row=total_row, column=4 + i).value
        per_reel_totals[f"reel{i}"] = round(float(v), 4) if isinstance(v, (int, float)) else None

    # Paytable starts a few rows below; look for "Combination" header in column C
    pay_header_row = None
    for r in range(total_row, total_row + 8):
        if ws.cell(row=r, column=3).value == 'Combination':
            pay_header_row = r
            break

    paytable = {}
    if pay_header_row is not None:
        # Walk paytable rows
        r = pay_header_row + 1
        while r < pay_header_row + 100:
            combo = [ws.cell(row=r, column=3 + i).value for i in range(5)]
            pay = ws.cell(row=r, column=8).value
            if not any(c for c in combo) or pay in (None, ''):
                break
            # Determine match count + symbol from combo (stop at '--')
            match_sym = combo[0]
            match_count = 0
            if match_sym in (None, '--'):
                r += 1
                continue
            for c in combo:
                if c == match_sym or c == 'Wild':
                    match_count += 1
                else:
                    break
            if isinstance(pay, (int, float)) and match_sym and match_count >= 3:
                paytable.setdefault(match_sym, {})[str(match_count)] = int(pay)
            r += 1

    return {
        "source": str(xlsx_path),
        "sheet": sheet_name,
        "title": title,
        "topology": topology,
        "swid": swid_row,
        "symbols": symbols,
        "per_reel_weights": per_reel_weights,
        "per_reel_totals": per_reel_totals,
        "paytable": paytable,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--sheet', default='PAR-001')
    ap.add_argument('--out', required=True)
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"▸ xlsx missing: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    data = extract(xlsx_path, args.sheet)

    # UQ-DEEP-M fix (Boki 2026-06-23): support --out '-' za stdout mode
    # (matches uniformni CLI shape sa par-sheet-pragmatic.py / lw.py).
    # Stdout MORA biti čist JSON; sve progress lines idu na stderr.
    # Plus: NORMALIZE shape u canonical ParBlob:
    #   per_reel_weights: 'reel0' keys → '0' numeric strings (bridge expect)
    #   paytable: dict {sym: combos} → list [{symbolId, combos}] (bridge expect)
    if args.out == '-' or not args.out:
        canonical = {**data}
        # Strip 'reel' prefix from per_reel_weights keys.
        if isinstance(data.get('per_reel_weights'), dict):
            canonical['per_reel_weights'] = {
                k.replace('reel', '') if k.startswith('reel') else k: v
                for k, v in data['per_reel_weights'].items()
            }
        # Convert paytable dict → list of {symbolId, combos}.
        if isinstance(data.get('paytable'), dict):
            canonical['paytable'] = [
                {'symbolId': sym, 'combos': combos}
                for sym, combos in data['paytable'].items()
                if combos
            ]
        sys.stdout.write(json.dumps(canonical) + '\n')
        print(f"✓ extracted {len(data['symbols'])} symbols × 5 reels", file=sys.stderr)
        return

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, 'w') as f:
        json.dump(data, f, indent=2)

    print(f"✓ extracted {len(data['symbols'])} symbols × 5 reels + {sum(len(v) for v in data['paytable'].values())} paytable rows", file=sys.stderr)
    print(f"  Title: {data['title']}", file=sys.stderr)
    print(f"  SWID:  {data['swid']}", file=sys.stderr)
    print(f"  Out:   {out_path}", file=sys.stderr)


if __name__ == '__main__':
    main()
